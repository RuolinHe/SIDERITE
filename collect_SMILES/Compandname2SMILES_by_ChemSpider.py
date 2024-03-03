import openpyxl
from chemspipy import ChemSpider
cs = ChemSpider('xxx') # here is your ChemSpider API
workbook=openpyxl.load_workbook("Sid_structure_output.xlsx")
worksheet=workbook["1-2"]
rows=worksheet.max_row
for i in range(2, rows+1):
    try:
        smiles = worksheet.cell(row=i, column=9).value
        if smiles is None:
            c1 = cs.search(worksheet.cell(row=i, column=3).value)
            if c1.count > 0:
                info=cs.get_details(c1[0].record_id)
                worksheet.cell(i,9,info['smiles'])
                formula = info['formula']
                formula = formula.replace("_{","")
                formula = formula.replace("}","")
                worksheet.cell(i,11,)
                worksheet.cell(i,14,info['commonName'])
                worksheet.cell(i,13,info['molecularWeight'])
    except:
        continue
workbook.save(filename="Sid_structure_output2.xlsx")
