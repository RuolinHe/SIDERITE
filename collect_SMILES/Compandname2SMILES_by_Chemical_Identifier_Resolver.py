import cirpy
import openpyxl
workbook=openpyxl.load_workbook("Sid_structure.xlsx")
worksheet=workbook["1-2"]
rows=worksheet.max_row
for i in range(2, rows+1):
    try:
        worksheet.cell(i,9,cirpy.resolve(worksheet.cell(row=i, column=3).value, 'smiles'))
        worksheet.cell(i,11,cirpy.resolve(worksheet.cell(row=i, column=3).value, 'formula'))
        worksheet.cell(i,12,cirpy.resolve(worksheet.cell(row=i, column=3).value, 'iupac_name'))
        worksheet.cell(i,13,cirpy.resolve(worksheet.cell(row=i, column=3).value, 'mw'))
    except:
        continue
workbook.save(filename="Sid_structure_output.xlsx")